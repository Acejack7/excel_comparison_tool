#!python 3

import os
import xlsxwriter
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# align language codes with target languages
languages = {'de-de': ['german', 'deutsch'],
             'es-es': ['spanish', 'español'],
             'fr-ca': ['french (canada), french, français'],
             'fr-fr': ['french', 'français'],
             'ja-jp': ['japanese', '日本人'],
             'it-it': ['italian', 'イタリアの'],
             'pt-br': ['portugese', 'portugese (brazilian)'],
             'pl-pl': ['polish', 'polski'],
             'sv-se': ['swedish', 'svenska'],
             }


# verify if valid column letter
def verify_column(col):
    if (len(col)) != 1:
        return(False)

    else:
        col_pattern = re.compile('[a-zA-Z]')
        col_check = re.findall(col_pattern, col)
        if len(col_check) == 1:
            return(True)
        else:
            return(False)


# check if input path is a file or dir; return the file list, dir of files and lang code
def get_files(file_path):
    if os.path.isdir(file_path):
        file_path_split = file_path.split('\\')
        cur_dir = ''
        for elem in file_path_split[:-1]:
            cur_dir += elem
            if file_path_split != elem:
                cur_dir += '\\'

        file_list = os.listdir(file_path)
        full_file_list = []
        for f in file_list:
            new_f = os.path.join(file_path, f)
            full_file_list.append(new_f)

        lang_dir = file_path_split[-1]
        if len(lang_dir) == 4:
            lang_code = lang_dir[:2] + '-' + lang_dir[2:]
        elif len(lang_dir) == 5:
            lang_code = lang_dir[:2] + '-' + lang_dir[3:]
        else:
            lang_code = ''

        return({'files': full_file_list, 'cur_dir': cur_dir, 'lang_code': lang_code})

    elif os.path.isfile(file_path):
        file_path_split = file_path.split('\\')
        file_path_split = file_path_split[:-1]
        cur_dir = ''
        for elem in file_path_split:
            cur_dir += elem
            if file_path_split != elem:
                cur_dir += '\\'

        lang_dir = file_path_split[-1]
        if len(lang_dir) == 4:
            lang_code = lang_dir[:2] + '-' + lang_dir[2:]
        elif len(lang_dir) == 5:
            lang_code = lang_dir[:2] + '-' + lang_dir[3:]
        else:
            lang_code = ''

        return({'files': [file_path], 'cur_dir': cur_dir, 'lang_code': lang_code})


# try to recognize target language by aligning lang code with dict languages
def get_target_lang(lang_code):
    if lang_code in languages:
        return(languages[lang_code])
    else:
        return('')


# verify if file is an excel and get rid of files that cannot be opened by load_workbook method
def verify_excel(files):
    proper_excel_files = []

    for file in files:
        try:
            wb = load_workbook(file)
            wb.close()
            proper_excel_files.append(file)
        except Exception:
            print('%s is not an excel file and got removed from the list.' % file)
            continue

    return(proper_excel_files)


# get excel contents: source, target, file, sheet and row of source/target
def get_excel_contents(files, target_lang, user_src_col='', user_trg_col=''):

    segments = []

    for file_elem in files:
        wb = load_workbook(file_elem)

        ws = wb.worksheets

        for sheet in ws:
            source_col = user_src_col
            target_col = user_trg_col

            if sheet.sheet_state == 'visible':
                start_row = 0

                if source_col == '':
                    for row in sheet.rows:
                        for cell in row:
                            cell_value = str(cell.value).lower()
                            if cell_value.startswith('english') or cell_value.startswith('source'):
                                start_row = cell.row
                                source_col = cell.column
                                break
                        if source_col != '':
                            break

                if target_col == '':
                    for col in sheet.iter_cols(min_row=start_row, max_row=start_row):
                        for cell in col:
                            cell_value = str(cell.value).lower()
                            if cell_value.startswith('translation') or cell_value.startswith('target'):
                                target_col = cell.column
                                break
                            elif target_lang != '':
                                if cell_value in target_lang:
                                    target_col = cell.column
                                    break
                        if target_col != '':
                            break

                if source_col != '' and target_col != '':
                    for num in range(start_row+1, sheet.max_row):
                        source_content = sheet[source_col + str(num)].value
                        target_content = sheet[target_col + str(num)].value
                        if source_content is not None and target_content is not None:
                            segments.append({'source': source_content, 'target': target_content,
                                             'file': file_elem, 'sheet': sheet.title, 'row': str(num)})

    return(segments)


# compare reviewed and translated content and align them
def compare_contents(translated_content, reviewed_content):
    full_content = []

    for translation in translated_content:
        for review in reviewed_content:
            if translation['source'] == review['source']:
                source_seg = translation['source']
                target_seg = translation['target']
                rev_seg = review['target']

                trans_file = os.path.split(translation['file'])[1]
                trans_sheet = translation['sheet']
                trans_row = translation['row']

                rev_file = os.path.split(review['file'])[1]
                rev_sheet = review['sheet']
                rev_row = review['row']
                full_content.append({'source': source_seg, 'target': target_seg, 'review': rev_seg,
                                     'trans_file': trans_file, 'trans_sheet': trans_sheet, 'trans_row': trans_row,
                                     'rev_file': rev_file, 'rev_sheet': rev_sheet, 'rev_row': rev_row})
                reviewed_content.remove(review)

    return(full_content)


# move to beginning segments that translation is different than review
def sort_by_changes(full_content):
    sorted_full_content = []

    for elem in full_content:
        if elem['target'] != elem['review']:
            sorted_full_content.insert(0, elem)
            elem['changed'] = True
        else:
            sorted_full_content.append(elem)
            elem['changed'] = False

    return(sorted_full_content)


# find different point (mark) between translated and reviewed content and save it as list
def mark_changes_in_rev(full_content):
    full_content_marked = []

    for elem in full_content:
        if elem['changed'] is True:
            trans = elem['target']
            rev = elem['review']

            trans_split = trans.split(' ')
            rev_split = rev.split(' ')

            review_text = []
            rev_same = ''
            rev_diff = ''

            for index, trans_elem in enumerate(trans_split):
                # catch exception where len(trans_split) > len(rev_split)
                try:
                    rev_elem = rev_split[index]
                except IndexError:
                    break

                if trans_elem == rev_elem:
                    rev_same += rev_elem + ' '
                else:
                    if len(rev_split) > len(trans_split):
                        for dif_txt in rev_split[index:len(trans_split)]:
                            rev_diff += dif_txt + ' '
                    else:
                        for dif_txt in rev_split[index:]:
                            rev_diff += dif_txt + ' '
                    break

            review_text.append(rev_same)
            review_text.append(rev_diff)

            # check if rev_split > trans_split to add the rest of the rev text to last element of review_text
            if len(rev_split) > len(trans_split):
                len_diff = len(rev_split) - len(trans_split)
                missed_words = rev_split[-len_diff:]
                for word in missed_words:
                    rev_diff += word + ' '
                review_text[1] = rev_diff

            # remove last redundant space
            review_text[1] = review_text[1].rstrip(' ')

            # assign the review_text list as review text
            elem['review'] = review_text

        full_content_marked.append(elem)

    return(full_content_marked)


# create reprot file with all information about translation and review
def create_report_file(full_content_marked, cur_dir, lang_code):
    report_file_path = os.path.join(cur_dir, lang_code + '_report.xlsx')

    wb_report = xlsxwriter.Workbook(report_file_path)
    ws = wb_report.add_worksheet('report')

    color_red = wb_report.add_format({'font_color': 'red'})
    color_green = wb_report.add_format({'font_color': 'green'})

    text_wrap = wb_report.add_format()
    text_wrap.set_text_wrap()

    red_color_wrap = wb_report.add_format()
    red_color_wrap.set_text_wrap()
    red_color_wrap.set_font_color('red')

    red_bg = wb_report.add_format()
    red_bg.set_pattern(1)
    red_bg.set_bg_color('red')

    ws.write('A1', 'Source')
    ws.write('B1', 'Translation')
    ws.write('C1', 'Review')
    ws.write('D1', 'Changed?')
    ws.write('E1', 'Trans File')
    ws.write('F1', 'Trans Sheet')
    ws.write('G1', 'Trans Row')
    ws.write('H1', 'Review File')
    ws.write('I1', 'Review Sheet')
    ws.write('J1', 'Review Row')

    ws.set_column("A:C", 35)
    ws.set_column("D:D", 10)
    ws.set_column("E:E", 50)
    ws.set_column("F:F", 20)
    ws.set_column("G:G", 15)
    ws.set_column("H:H", 50)
    ws.set_column("I:I", 20)
    ws.set_column("J:J", 15)

    row_num = 2
    for seg in full_content_marked:
        cur_row = str(row_num)

        source = seg['source']
        target = seg['target']
        rev = seg['review']

        trans_file = seg['trans_file']
        trans_sheet = seg['trans_sheet']
        trans_row = seg['trans_row']

        review_file = seg['rev_file']
        review_sheet = seg['rev_sheet']
        review_row = seg['rev_row']

        ws.write('A' + cur_row, source, text_wrap)
        ws.write('B' + cur_row, target, text_wrap)
        if type(rev) is str:
            ws.write('C' + cur_row, rev, text_wrap)
        else:
            if rev[0] == '' or rev[1] == '':
                ws.write('C' + cur_row, rev[0] + rev[1], red_color_wrap)
            else:
                ws.write_rich_string('C' + cur_row, color_green, rev[0], color_red, rev[1], text_wrap)

        if seg['changed'] is False:
            ws.write('D' + cur_row, 'No')
        else:
            ws.write('D' + cur_row, 'Yes', red_bg)

        ws.write('E' + cur_row, trans_file)
        ws.write('F' + cur_row, trans_sheet)
        ws.write('G' + cur_row, trans_row)

        ws.write('H' + cur_row, review_file)
        ws.write('I' + cur_row, review_sheet)
        ws.write('J' + cur_row, review_row)

        row_num += 1

    wb_report.close()
    return('Report file created successfully.')


# for now (24.02.2019) outdated menthod of creating/updating report file
def add_data_to_report(full_content, report_file):
    wb_report = load_workbook(report_file)

    ws = wb_report.worksheets[0]

    ws['A1'] = 'Source'
    ws['B1'] = 'Translation'
    ws['D1'] = 'Changed?'
    ws['E1'] = 'Trans File'
    ws['F1'] = 'Trans Sheet'
    ws['G1'] = 'Trans Row'
    ws['H1'] = 'Review File'
    ws['I1'] = 'Review Sheet'
    ws['J1'] = 'Review Row'

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 15

    counter = 2
    for content in full_content:
        row = str(counter)
        source = content['source']
        target = content['target']

        trans_file = content['trans_file']
        trans_sheet = content['trans_sheet']
        trans_row = content['trans_row']

        review_file = content['rev_file']
        review_sheet = content['rev_sheet']
        review_row = content['rev_row']

        ws['A' + row] = source
        ws['A' + row].alignment = Alignment(wrap_text=True)
        ws['B' + row] = target
        ws['B' + row].alignment = Alignment(wrap_text=True)
        ws['C' + row].alignment = Alignment(wrap_text=True)

        if content['changed'] is False:
            ws['D' + row] = 'No'
        else:
            ws['D' + row] = 'Yes'
            ws['D' + row].fill = PatternFill(fgColor='FF0000', fill_type='solid')

        ws['E' + row] = trans_file
        ws['F' + row] = trans_sheet
        ws['G' + row] = trans_row

        ws['H' + row] = review_file
        ws['I' + row] = review_sheet
        ws['J' + row] = review_row

        counter += 1

    wb_report.save(report_file)
    return('Report file is updated.')
